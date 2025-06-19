from django.utils.timezone import timedelta, now
from openpyxl import Workbook
from dominoapp.connectors.google_verifier import GoogleDrive
from dominoapp.connectors.discord_connector import DiscordConnector
import tempfile
import os
import logging  
logger = logging.getLogger('django')

def create_xlsx_file(title:str, players):
    if title is None:
        raise('some requirement is lost')
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = Workbook()
        ws = wb.active
        ws.title = title

        Time_in_Cuba = now() - timedelta(hours = 4)
        ws['A1'] = Time_in_Cuba.strftime("%d-%m-%Y %H:%M:%S")
        ws['A2'] = 'Alias'
        ws['B2'] = 'Erned Coins'
        ws['C2'] = 'Recharged Coins'
        ws['D2'] = 'Email'
        ws['E2'] = 'Name'
        ws['F2'] = 'Last Time System'

        for index, player in enumerate(players, start=3):
            ws[f'A{index}'] = player.alias
            ws[f'B{index}'] = player.earned_coins
            ws[f'C{index}'] = player.recharged_coins
            ws[f'D{index}'] = player.email
            ws[f'E{index}'] = player.name
            ws[f'F{index}'] = player.lastTimeInSystem.strftime("%d-%m-%Y %H:%M:%S")

        wb.save(tmp.name)
        
        # Subir a Google Drive
        try:
            result = GoogleDrive.upload_file(
                file_path=tmp.name,
                file_name='player_backup.xlsx'
            )
            
            DiscordConnector.send_event(
                "Nueva Copia de Seguridad",
                result['webViewLink']
            )
            GoogleDrive.delete_old_file("player_backup.xlsx", result['id']) 
        except Exception as e:
            logger.critical(f'No se completo la exportacion de los player. Error => {str(e)}')
          
        finally:
            # Eliminar archivo temporal
            tmp.close()
            os.unlink(tmp.name)

    return